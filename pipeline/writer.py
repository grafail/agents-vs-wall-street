"""Workbook writer: fill ONLY the three yellow forecast cells of a supplied
template, refusing loudly on any structural mismatch.

Template layout (verified identical across all 4 templates on 2026-08-16):
    sheet "Summary"
    A6:C6   header row ("Metric" | "Units" | <period, e.g. FY2026Q2>)
    A7:A9   metric labels (verbatim contest labels)
    B7:B9   unit strings  (verbatim contest units)
    C7:C9   yellow forecast cells (fill FFFFF7D6) — the ONLY cells we write
"""
from __future__ import annotations

import math
import shutil
import subprocess
from pathlib import Path

import openpyxl

from pipeline.config import ROOT, SUBMISSION_DIR, TEMPLATES_DIR
from pipeline.types import MetricSpec

SHEET_NAME = "Summary"
PERIOD_CELL = "C6"
FIRST_METRIC_ROW = 7
VALUE_COL = "C"
LABEL_COL = "A"
UNIT_COL = "B"


class TemplateMismatch(RuntimeError):
    """Raised when the template does not match the contest spec verbatim."""


def _check_structure(ws_title_ok: bool, msg: str) -> None:
    if not ws_title_ok:
        raise TemplateMismatch(msg)


def write_workbook(
    spec_group: list[MetricSpec],
    values: dict[str, float],
    out_dir: Path = SUBMISSION_DIR,
) -> Path:
    """Fill the three value cells for one company's template and save it under
    the exact contest output_file name. Refuses (raises) on ANY mismatch
    between the template and the typed spec — labels, units, period, sheet name
    — and on any non-finite / non-numeric value. Never touches the template.
    """
    if not spec_group:
        raise TemplateMismatch("empty spec_group")
    output_files = {s.output_file for s in spec_group}
    _check_structure(len(output_files) == 1,
                     f"spec_group spans multiple output files: {output_files}")
    output_file = spec_group[0].output_file

    template_path = TEMPLATES_DIR / output_file
    _check_structure(template_path.exists(), f"template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    _check_structure(SHEET_NAME in wb.sheetnames,
                     f"{output_file}: no sheet named {SHEET_NAME!r} (found {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    period = ws[PERIOD_CELL].value
    for spec in spec_group:
        _check_structure(
            period == spec.period,
            f"{output_file}: period header {PERIOD_CELL}={period!r} != spec.period {spec.period!r}")

    for i, spec in enumerate(spec_group):
        row = FIRST_METRIC_ROW + i
        label = ws[f"{LABEL_COL}{row}"].value
        unit = ws[f"{UNIT_COL}{row}"].value
        _check_structure(
            label == spec.label,
            f"{output_file}: label cell {LABEL_COL}{row}={label!r} != spec.label {spec.label!r}")
        _check_structure(
            unit == spec.unit_str,
            f"{output_file}: unit cell {UNIT_COL}{row}={unit!r} != spec.unit_str {spec.unit_str!r}")

        if spec.label not in values:
            raise TemplateMismatch(f"{output_file}: no value provided for {spec.label!r} "
                                   "(never-blank rule: caller must supply a number)")
        raw = values[spec.label]
        if isinstance(raw, bool) or not isinstance(raw, (int, float)):
            raise TemplateMismatch(
                f"{output_file}: value for {spec.label!r} is {type(raw).__name__}, not a number")
        value = float(raw)
        if not math.isfinite(value):
            raise TemplateMismatch(f"{output_file}: value for {spec.label!r} is not finite: {value}")

        ws[f"{VALUE_COL}{row}"] = value  # written as a number, never a string

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / output_file
    wb.save(out_path)
    return out_path


def verify_workbook(path: Path, specs: list[MetricSpec]) -> list[str]:
    """Reopen a written workbook and re-check everything: sheet name, verbatim
    labels/units/period, and that every value cell holds a finite number.
    Returns a list of issue strings (empty == clean). Never raises on content
    problems — only on an unreadable file.
    """
    issues: list[str] = []
    path = Path(path)
    if not path.exists():
        return [f"file missing: {path}"]

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        return [f"{path.name}: no sheet named {SHEET_NAME!r} (found {wb.sheetnames})"]
    ws = wb[SHEET_NAME]

    for i, spec in enumerate(specs):
        row = FIRST_METRIC_ROW + i
        label = ws[f"{LABEL_COL}{row}"].value
        unit = ws[f"{UNIT_COL}{row}"].value
        period = ws[PERIOD_CELL].value
        value = ws[f"{VALUE_COL}{row}"].value

        if label != spec.label:
            issues.append(f"{path.name} {LABEL_COL}{row}: label {label!r} != {spec.label!r}")
        if unit != spec.unit_str:
            issues.append(f"{path.name} {UNIT_COL}{row}: unit {unit!r} != {spec.unit_str!r}")
        if period != spec.period:
            issues.append(f"{path.name} {PERIOD_CELL}: period {period!r} != {spec.period!r}")

        if value is None:
            issues.append(f"{path.name} {VALUE_COL}{row}: forecast is blank ({spec.label})")
        elif isinstance(value, bool) or not isinstance(value, (int, float)):
            issues.append(f"{path.name} {VALUE_COL}{row}: forecast is {type(value).__name__} "
                          f"{value!r}, not a number ({spec.label})")
        elif not math.isfinite(float(value)):
            issues.append(f"{path.name} {VALUE_COL}{row}: forecast not finite: {value} ({spec.label})")
    return issues


def run_submission_check() -> tuple[bool | None, str]:
    """Shell out to `npm run check:submission` from the repo root.
    Returns (ok, output); (None, "npm not found") when npm is unavailable.
    """
    if shutil.which("npm") is None:
        return None, "npm not found"
    try:
        proc = subprocess.run(
            ["npm", "run", "check:submission"],
            cwd=ROOT, capture_output=True, text=True, timeout=120,
        )
    except (subprocess.TimeoutExpired, OSError) as exc:
        return None, f"npm run check:submission failed to run: {exc}"
    return proc.returncode == 0, (proc.stdout + proc.stderr)
