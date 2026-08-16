"""Workbook writer against the four REAL templates. Writes only to tmp_path;
asserts the originals are never modified."""
import hashlib
import math

import openpyxl
import pytest

from pipeline.config import TEMPLATES_DIR
from pipeline.types import MetricSpec, metric_specs
from pipeline.writer import (
    TemplateMismatch,
    run_submission_check,
    verify_workbook,
    write_workbook,
)

# Plausible dummy forecasts, keyed by exact contest label per file.
VALUES = {
    "HD-FY2026Q2.xlsx": {
        "Net sales": 45_300.0,
        "Adjusted diluted EPS": 4.68,
        "Comparable sales, total company": 1.5,
    },
    "ADI-FY2026Q3.xlsx": {
        "Revenue": 2_900.0,
        "Adjusted diluted EPS": 2.05,
        "Adjusted gross margin": 69.5,
    },
    "HAS-FY2026.xlsx": {
        "Net fees": 1_050.0,
        "Pre-exceptional basic EPS": 6.2,
        "Pre-exceptional operating profit": 60.0,
    },
    "DE-FY2026Q3.xlsx": {
        "Worldwide net sales and revenues": 12_100.0,
        "Diluted EPS (GAAP)": 5.20,
        "Production & Precision Ag operating profit": 2_250.0,
    },
}


def spec_groups() -> dict[str, list[MetricSpec]]:
    groups: dict[str, list[MetricSpec]] = {}
    for s in metric_specs():
        groups.setdefault(s.output_file, []).append(s)
    return groups


def _sha(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.fixture()
def template_hashes():
    return {p.name: _sha(p) for p in TEMPLATES_DIR.glob("*.xlsx")}


def test_all_four_templates_and_specs_line_up():
    groups = spec_groups()
    assert set(groups) == set(VALUES)
    for name, specs in groups.items():
        assert len(specs) == 3
        assert (TEMPLATES_DIR / name).exists()


@pytest.mark.parametrize("output_file", sorted(VALUES))
def test_write_and_verify_each_template(output_file, tmp_path, template_hashes):
    specs = spec_groups()[output_file]
    out = write_workbook(specs, VALUES[output_file], out_dir=tmp_path)

    assert out == tmp_path / output_file
    assert out.exists()

    # verify_workbook is clean on a good write
    assert verify_workbook(out, specs) == []

    # value cells are numeric floats, structure untouched
    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    assert ws["C6"].value == specs[0].period
    for i, spec in enumerate(specs):
        row = 7 + i
        assert ws[f"A{row}"].value == spec.label
        assert ws[f"B{row}"].value == spec.unit_str
        v = ws[f"C{row}"].value
        assert isinstance(v, (int, float)) and not isinstance(v, (bool, str))
        assert math.isfinite(float(v))
        assert float(v) == pytest.approx(VALUES[output_file][spec.label])

    # original template NEVER modified
    assert _sha(TEMPLATES_DIR / output_file) == template_hashes[output_file]


def test_refuses_wrong_label(tmp_path):
    specs = [s.model_copy(deep=True) for s in spec_groups()["HD-FY2026Q2.xlsx"]]
    specs[0].label = "Net revenue"  # tampered spec != template cell A7
    values = dict(VALUES["HD-FY2026Q2.xlsx"])
    values["Net revenue"] = values.pop("Net sales")
    with pytest.raises(TemplateMismatch, match="label"):
        write_workbook(specs, values, out_dir=tmp_path)
    assert not (tmp_path / "HD-FY2026Q2.xlsx").exists()


def test_refuses_wrong_unit(tmp_path):
    specs = [s.model_copy(deep=True) for s in spec_groups()["ADI-FY2026Q3.xlsx"]]
    specs[1].unit_str = "GBp"
    with pytest.raises(TemplateMismatch, match="unit"):
        write_workbook(specs, VALUES["ADI-FY2026Q3.xlsx"], out_dir=tmp_path)


def test_refuses_wrong_period(tmp_path):
    specs = [s.model_copy(deep=True) for s in spec_groups()["DE-FY2026Q3.xlsx"]]
    for s in specs:
        s.period = "FY2027Q1"
    with pytest.raises(TemplateMismatch, match="period"):
        write_workbook(specs, VALUES["DE-FY2026Q3.xlsx"], out_dir=tmp_path)


def test_refuses_non_finite_and_non_numeric_values(tmp_path):
    specs = spec_groups()["HAS-FY2026.xlsx"]
    bad = dict(VALUES["HAS-FY2026.xlsx"])
    bad["Net fees"] = float("nan")
    with pytest.raises(TemplateMismatch, match="finite"):
        write_workbook(specs, bad, out_dir=tmp_path)
    bad["Net fees"] = "1050"  # strings must be refused, never coerced
    with pytest.raises(TemplateMismatch, match="not a number"):
        write_workbook(specs, bad, out_dir=tmp_path)
    bad["Net fees"] = True  # bools are ints in Python; still refuse
    with pytest.raises(TemplateMismatch, match="not a number"):
        write_workbook(specs, bad, out_dir=tmp_path)


def test_refuses_missing_value(tmp_path):
    specs = spec_groups()["HD-FY2026Q2.xlsx"]
    values = dict(VALUES["HD-FY2026Q2.xlsx"])
    del values["Net sales"]
    with pytest.raises(TemplateMismatch, match="no value"):
        write_workbook(specs, values, out_dir=tmp_path)


def test_refuses_mixed_spec_group(tmp_path):
    groups = spec_groups()
    mixed = groups["HD-FY2026Q2.xlsx"][:2] + groups["ADI-FY2026Q3.xlsx"][:1]
    with pytest.raises(TemplateMismatch, match="multiple output files"):
        write_workbook(mixed, VALUES["HD-FY2026Q2.xlsx"], out_dir=tmp_path)


def test_verify_detects_blank_and_tampered_cells(tmp_path):
    specs = spec_groups()["HD-FY2026Q2.xlsx"]
    out = write_workbook(specs, VALUES["HD-FY2026Q2.xlsx"], out_dir=tmp_path)

    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    ws["C7"] = None          # blank a forecast
    ws["C8"] = "4.68"        # string, not a number
    ws["A9"] = "Something else"  # tamper a label
    wb.save(out)

    issues = verify_workbook(out, specs)
    assert any("blank" in i for i in issues)
    assert any("not a number" in i for i in issues)
    assert any("label" in i for i in issues)


def test_verify_missing_file():
    specs = spec_groups()["HD-FY2026Q2.xlsx"]
    issues = verify_workbook(TEMPLATES_DIR / "nope.xlsx", specs)
    assert issues and "missing" in issues[0]


def test_run_submission_check_tolerates_missing_npm(monkeypatch):
    import pipeline.writer as writer_mod
    monkeypatch.setattr(writer_mod.shutil, "which", lambda _: None)
    ok, output = run_submission_check()
    assert ok is None
    assert output == "npm not found"
